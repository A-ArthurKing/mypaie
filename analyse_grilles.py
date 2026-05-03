import openpyxl, os

folder = 'docs/grilles_sources'
files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and 'LANGUE' not in f]

output = []
for fname in sorted(files):
    path = os.path.join(folder, fname)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        output.append(f'\n========== {fname} ==========')
        for sheet_name in ['Variable', 'OBJECTIFS']:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            output.append(f'  -- {sheet_name} --')
            for row in ws.iter_rows(min_row=1, max_row=100, max_col=15):
                for cell in row:
                    if cell.value is not None:
                        output.append(f'    [{cell.coordinate}] {repr(cell.value)}')
    except Exception as e:
        output.append(f'  ERREUR: {e}')

with open('docs/grilles_sources/_analyse.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(output))
print('OK - fichier ecrit')
